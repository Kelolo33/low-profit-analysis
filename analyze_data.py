import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import os

def process_subscription_file(subscription_file):
    # 读取海运订阅文件，移除 encoding 参数
    df = pd.read_excel(subscription_file)
    
    print("海运订阅文件的列名:", df.columns.tolist(), flush=True)
    print(f"原始数据行数: {len(df)}", flush=True)
    
    # 检查必需的列
    required_columns = ['二级部门', '委托客户', '客户约价', '是否低负', '未税人民币总毛利', '未税人民币总收入', '业务大类名称', '业务月度']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"海运订阅文件缺少以下列: {', '.join(missing_columns)}")
    
    # 获取业务月度并进行验证
    business_month = None
    if not df.empty:
        # 获取非空的业务月度值
        valid_months = df['业务月度'].dropna()
        if not valid_months.empty:
            # 获取第一个非空值
            first_month = valid_months.iloc[0]
            if pd.notna(first_month) and str(first_month) != 'nan':
                business_month = str(first_month)
    
    # 如果没有有效的业务月度，使用当前日期
    if not business_month:
        from datetime import datetime
        business_month = datetime.now().strftime("%Y-%m")
        print(f"警告：未找到有效的业务月度，使用当前日期：{business_month}")
    
    # 筛选业务大类为海运的数据
    df = df[df['业务大类名称'] == '海运']
    print(f"筛选海运业务后的数据行数: {len(df)}", flush=True)
    
    # 选择所需列
    subscription_data = df[required_columns].copy()
    print(f"选择所需列后的数据行数: {len(subscription_data)}", flush=True)
    
    # 计算约价的负毛利票数和非约价的低负票数
    subscription_data['约价负毛利'] = ((subscription_data['客户约价'].notna() & (subscription_data['客户约价'] != 'N')) & 
                                  (subscription_data['是否低负'] == '负毛利'))
    subscription_data['非约价低负'] = ((subscription_data['客户约价'].isna() | (subscription_data['客户约价'] == 'N')) & 
                                  ((subscription_data['是否低负'] == '低毛利') | (subscription_data['是否低负'] == '负毛利')))
    
    print(f"约价负毛利的记录数: {subscription_data['约价负毛利'].sum()}", flush=True)
    print(f"非约价低负的记录数: {subscription_data['非约价低负'].sum()}", flush=True)
    
    # 分别计算约价和非约价的数据，只考虑 '是否低' 不为 'N' 的数据
    yue_data = subscription_data[(subscription_data['约价负毛利']) & (subscription_data['是否低负'] != 'N')].groupby(['二级部门', '委托客户']).agg({
        '未税人民币总毛利': 'sum',
        '未税人民币总收入': 'sum',
        '约价负毛利': 'count'
    }).reset_index()
    
    non_yue_data = subscription_data[(subscription_data['非约价低负']) & (subscription_data['是否低负'] != 'N')].groupby(['二级部门', '委托客户']).agg({
        '未税人民币总毛利': 'sum',
        '未税人民币总收入': 'sum',
        '非约价低负': 'count'
    }).reset_index()
    
    print(f"约价数据行数: {len(yue_data)}")
    print(f"非约价数据行数: {len(non_yue_data)}")
    
    # 所有委托客户
    all_customers = df[['二级部门', '委托客户']].drop_duplicates()
    
    # 合并约价和非约价数据，并确保包含所有委托客户
    grouped_data = pd.merge(all_customers, yue_data, on=['二级部门', '委托客户'], how='left')
    grouped_data = pd.merge(grouped_data, non_yue_data, on=['二级部门', '委托客户'], how='left')
    
    # 重命名列
    grouped_data = grouped_data.rename(columns={
        '约价负毛利': '约价负毛利票数',
        '非约价低负': '非约价低负票数',
        '未税人民币总毛利_x': '约价未税人民币总毛利',
        '未税人民币总收入_x': '约价未税人民币总收入',
        '未税人民币总毛利_y': '非约价未税人民币总毛利',
        '未税人民币总收入_y': '非约价未税人民币总收入',
    })
    
    # 填充NaN值为0
    fill_columns = ['约价负毛利票数', '非约价低负票数', '约价未税人民币总毛利', '约价未税人民币总收入', '非约价未税人民币总毛利', '非约价未税人民币总收入']
    grouped_data[fill_columns] = grouped_data[fill_columns].fillna(0)
    
    # 过滤掉约价负毛利票数和非约价低负票数都为0的记录
    grouped_data = grouped_data[
        (grouped_data['约价负毛利票数'] > 0) | 
        (grouped_data['非约价低负票数'] > 0)
    ]
    
    # 重新计算毛利率
    grouped_data['约价毛利率'] = np.where(
        grouped_data['约价未税人民币总收入'] != 0,
        grouped_data['约价未税人民币总毛利'] / grouped_data['约价未税人民币总收入'],
        -1
    )
    grouped_data['非约价毛利率'] = np.where(
        grouped_data['非约价未税人民币总收入'] != 0,
        grouped_data['非约价未税人民币总毛利'] / grouped_data['非约价未税人民币总收入'],
        -1
    )

    # 将毛利率转换为百分比格式的数值，票数为0时显示为空
    def format_rate(rate, count):
        if pd.isnull(rate) or pd.isnull(count) or count == 0:
            return None
        elif rate == -1:
            return -1  # 返回 -1，表示 -100%
        else:
            return rate  # 保持原始的小数形式
    
    grouped_data['约价毛利率'] = grouped_data.apply(lambda row: format_rate(row['约价毛利率'], row['约价负毛利票数']), axis=1)
    grouped_data['非约价毛利率'] = grouped_data.apply(lambda row: format_rate(row['非约价毛利率'], row['非约价低负票数']), axis=1)
    
    grouped_data['约价负毛利票数'] = grouped_data['约价负毛利票数'].apply(lambda x: '' if x == 0 else int(x))
    grouped_data['非约价低负票数'] = grouped_data['非约价低负票数'].apply(lambda x: '' if x == 0 else int(x))
    
    print("grouped_data 的前几行:")
    print(grouped_data.head().to_string())
    
    # 打印一些统信息
    print(f"\n约价负毛利票数不为空的记录数: {grouped_data['约价负毛利票数'].astype(bool).sum()}")
    print(f"非约价低负票数不为空的记录数: {grouped_data['非约价低负票数'].astype(bool).sum()}")
    print(f"约价毛利率不为空的记录数: {grouped_data['约价毛利率'].astype(bool).sum()}")
    print(f"非约价毛利率不为空的记录数: {grouped_data['非约价毛利率'].astype(bool).sum()}")
    
    # 计算每个委托客户的总毛利率
    total_stats = df.groupby(['二级部门', '委托客户']).agg({
        '未税人民币总毛利': 'sum',
        '未税人民币总收入': 'sum'
    }).reset_index()
    
    # 安全地计算总利润率
    def calculate_profit_rate(row):
        if pd.isna(row['未税人民币总收入']) or row['未税人民币总收入'] == 0:
            return 0  # 或者返回其他默认值
        return row['未税人民币总毛利'] / row['未税人民币总收入']
    
    total_stats['总利润率'] = total_stats.apply(calculate_profit_rate, axis=1)
    
    # 过滤掉异常值（如果需要的话）
    total_stats['总利润率'] = total_stats['总利润率'].apply(
        lambda x: x if abs(x) <= 1 else (1 if x > 1 else -1)
    )
    
    # 合并总利润率到grouped_data
    grouped_data = pd.merge(grouped_data, total_stats[['二级部门', '委托客户', '总利润率']], 
                          on=['二级部门', '委托客户'], how='left')
    
    return grouped_data, business_month

def analyze_excel_data(input_file, output_file, subscription_file, status_callback=None):
    if status_callback:
        status_callback("开始读取海运订阅文件...")
    
    # 读取海运订阅文件
    subscription_df = pd.read_excel(subscription_file)
    
    if status_callback:
        status_callback("处理海运订阅数据...")
    subscription_data, business_month = process_subscription_file(subscription_file)

    # 修改输出文件名，添加总表标识和业务月度，同时保持原始路径
    output_dir = os.path.dirname(output_file)
    base_name = "分析结果_总表"
    output_file = os.path.join(output_dir, f"{base_name}_{business_month}.xlsx")

    if input_file:
        if status_callback:
            status_callback("读取预对账文件...")
        df = pd.read_excel(input_file)
        
        if status_callback:
            status_callback("分析数据中...")
        # 确保必要的列存在
        required_columns = ['法人部门', '委托客户', '别名', '应收应付', '本位币金额', '费率单号', '币种']
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"缺少必要的列: {col}")

        # 按法人部门、委托客户、费率单号和别名进行汇总
        grouped = df.groupby(['法人部门', '委托客户', '费率单号', '别名', '应收应付', '币种'])['本位币金额'].sum().unstack(level='应收应付').fillna(0)
        grouped = grouped.rename(columns={'应收': '应收金额', '应付': '应付金额'})
        grouped['费目利润'] = grouped['应收金额'] - grouped['应付金额']

        # 重置索引，使得所有列都变成普通列
        grouped = grouped.reset_index()

        # 先计算每个费率单号的总毛利和毛利率
        rate_totals = grouped.groupby('费率单号').agg({
            '应收金额': 'sum',
            '应付金额': 'sum'
        }).reset_index()
        
        rate_totals['单票毛利'] = rate_totals['应收金额'] - rate_totals['应付金额']
        rate_totals['单票毛利率'] = rate_totals.apply(
            lambda x: x['单票毛利'] / x['应收金额'] if x['应收金额'] != 0 else -1, 
            axis=1
        )
        
        # 只保留需要的列
        rate_totals = rate_totals[['费率单号', '单票毛利', '单票毛利率']]

        # 创建结果 DataFrame
        results = []
        seen_rate_nos = {}  # 用于跟踪已经见过的费率单号

        for _, row in grouped.iterrows():
            # 获取该费率单号的总计数据
            rate_total = rate_totals[rate_totals['费率单号'] == row['费率单号']].iloc[0]

            result = {
                '法人部门': row['法人部门'],
                '委托客户': row['委托客户'],
                '费率单号': row['费率单号'],
                '别名': row['别名'],
                '币种': row['币种'],
                '应收金额': row['应收金额'],
                '应付金额': row['应付金额'],
                '费目利润': row['费目利润'],
                '类型': '',
                '单票毛利': rate_total['单票毛利'],  # 使用费率单总毛利
                '单票毛利率': rate_total['单票毛利率']  # 使用费率单总毛利率
            }

            if row['应付金额'] > 0 and row['应收金额'] == 0:
                result['类型'] = '无应收'
            elif row['应收金额'] < row['应付金额']:
                result['类型'] = '倒挂'

            results.append(result)

        # 创建结果 DataFrame
        result_df = pd.DataFrame(results)
        
        # 创建客户公司分析数据
        customer_analysis = result_df.groupby(['法人部门', '委托客户']).agg({
            '费率单号': lambda x: len(x.unique()),  # 统计唯一费率单号的数量
            '费目利润': 'sum',  # 总金额
        }).reset_index()

        customer_analysis['初步分析'] = result_df.groupby(['法人部门', '委托客户']).apply(format_analysis).reset_index(drop=True)
        customer_analysis = customer_analysis.rename(columns={
            '费率单号': '总票数',
            '费目利润': '总金额',
        })
    else:
        # 如果没有预对账文件，创建一个空的customer_analysis DataFrame
        customer_analysis = pd.DataFrame(columns=['法人部门', '委托客户', '总票数', '总金额', '初步分析'])

    # 创建法人部门和二级部门的对应关系映射
    dept_mapping = {
        '内贸水运': '内贸',
        '外贸水运': '外贸'
    }
    
    # 添加对应的法人部门列
    subscription_data['法人部门'] = subscription_data['二级部门'].map(lambda x: dept_mapping.get(x, x))
    
    # 将海运订阅文件中的所有二级部门和委托客户信息合并到客户分析结果中
    full_analysis = pd.merge(subscription_data, customer_analysis, 
                           on=['法人部门', '委托客户'], 
                           how='left')
    
    # 填充NaN值
    full_analysis = full_analysis.fillna({'总票数': 0, '总金额': 0, '总利润率': 0, '初步分析': ''})

    # 对full_analysis进行排序
    full_analysis = full_analysis.sort_values(by=['二级部门', '委托客户'])

    # 保存结果到 Excel 文件
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 首先保存海运订阅文件的原始数据
        subscription_df.to_excel(writer, index=False, sheet_name='海运订阅原始数据')
        # 设置海运订阅原始数据sheet的冻结窗格
        writer.sheets['海运订阅原始数据'].freeze_panes = 'A2'
        
        if input_file:
            # 保存预对账原始数据
            df.to_excel(writer, index=False, sheet_name='预对账原始数据')
            # 设置预对账原始数据sheet的冻结窗格
            writer.sheets['预对账原始数据'].freeze_panes = 'A2'
            
            # 处理分析结果sheet，应用"只显示一次"的逻辑
            display_df = result_df.copy()
            display_df = display_df.sort_values(['法人部门', '委托客户', '费率单号'])
            
            # 创建一个布尔掩码，标记每个费率单号的第一次出现
            is_first = ~display_df['费率单号'].duplicated()
            
            # 将非第一次出现的记录的特定字段设置为空
            # 修改：分别处理字符串列和数值列
            display_df.loc[~is_first, ['委托客户', '费率单号']] = ''  # 字符串列
            display_df.loc[~is_first, ['单票毛利', '单票毛利率']] = np.nan  # 数值列用 NaN
            
            # 保存处理后的分析结果
            display_df.to_excel(writer, index=False, sheet_name='分析结果')
            # 设置分析结果sheet的冻结窗格
            result_sheet = writer.sheets['分析结果']
            result_sheet.freeze_panes = 'A2'
            
            # 设置原始数据sheet的列宽
            if input_file:
                # 设置分析结果sheet的格式
                result_sheet = writer.sheets['分析结果']
                
                # 先设置固定列宽
                result_sheet.column_dimensions['A'].width = 17  # 法人部门列
                result_sheet.column_dimensions['I'].width = 8   # 别名列
                
                # 设置单票毛利率为百分比格式
                for row in result_sheet.iter_rows(min_row=2):  # 从第2行开始（跳过表头）
                    if row[10].value:  # 第11列是单票毛利率（K列）
                        row[10].number_format = '0.00%'
                
                # 对其他列进行自适应宽度设置
                for column in result_sheet.columns:
                    column_letter = get_column_letter(column[0].column)
                    if column_letter not in ['A', 'I']:  # 跳过已设置的列
                        max_length = 0
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)  # 限制最大宽度为30
                        result_sheet.column_dimensions[column_letter].width = adjusted_width
                
                # 设置预对账原始数据sheet的列宽
                worksheet = writer.sheets['预对账原始数据']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 40)  # 限制最大宽度为40
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # 设置海运订阅原始数据sheet的列宽
            worksheet = writer.sheets['海运订阅原始数据']
            # 先设置B列的固定宽度
            worksheet.column_dimensions['B'].width = 9
            
            # 对其他列进行自适应宽度设置
            for column in worksheet.columns:
                column_letter = get_column_letter(column[0].column)
                if column_letter != 'B':  # 跳过B列
                    max_length = 0
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 40)  # 限制最大宽度为40
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        # 创建客户公司分析sheet
        workbook = writer.book
        analysis_sheet = workbook.create_sheet(title='客户公司分析')
        analysis_sheet.freeze_panes = 'A3'  # 因为有两行表头，所以从第3行开始
        
        # 设置列宽
        column_widths = {
            'A': 15,  # 二级部门
            'B': 30,  # 客户公司
            'C': 12,  # 约价负毛利票数
            'D': 10,  # 约价毛利率
            'E': 12,  # 非约价低负票数
            'F': 10,  # 非约价毛利率
            'G': 10,  # 总票数
            'H': 10,  # 总利润率
            'I': 40,  # 初步分析
            'J': 40,  # 业务部门反馈具体原因
            'K': 15,  # 原因类别
            'L': 12,  # 损调利润
            'M': 40,  # 计划采取的措施
            'N': 15,  # 是否完成价格备案表
            'O': 15,  # 是否联合磋商
            'P': 15,  # 督办任务
            'Q': 10,  # 责任人
            'R': 15,  # 督办时间点
        }

        # 设置列宽
        for col, width in column_widths.items():
            analysis_sheet.column_dimensions[col].width = width

        # 添加表头
        headers = [
            ['二级部门', '委托客户', '约价', '约价', '非约价', '非约价', '总票数', '总利润率', '初步分析', '业务部门反馈具体原因', 
             '原因类别', '损调利润', '计划采取的措施', '是否完成价格备案表', '是否联合磋商', '督办任务', 
             '责任人', '督办时间点'],
            ['', '', '负毛利票数', '毛利率', '低负票数', '毛利率', '', '', '', '', '', '', '', '', '', '', '', '']
        ]
        
        # 设置表头样式
        for row_index, row in enumerate(headers, start=1):
            for col_index, value in enumerate(row, start=1):
                cell = analysis_sheet.cell(row=row_index, column=col_index, value=value)
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # 合并单元格
        merge_ranges = [
            'A1:A2',  # 二级部门
            'B1:B2',  # 委托客户
            'C1:D1',  # 约价
            'E1:F1',  # 非约价
            'G1:G2',  # 总票数
            'H1:H2',  # 总利润率
            'I1:I2',  # 初步分析
            'J1:J2',  # 业务部门反馈具体原因
            'K1:K2',  # 原因类别
            'L1:L2',  # 损调利润
            'M1:M2',  # 计划采取的措施
            'N1:N2',  # 是否完成价格备案表
            'O1:O2',  # 是否联合磋商
            'P1:P2',  # 督办任务
            'Q1:Q2',  # 责任人
            'R1:R2',  # 督办时间点
        ]

        # 执行单元格合并
        for cell_range in merge_ranges:
            analysis_sheet.merge_cells(cell_range)

        # 设置表头行高
        header_height = min(30, 50)  # 表头行高上限为50
        analysis_sheet.row_dimensions[1].height = header_height
        analysis_sheet.row_dimensions[2].height = header_height

        # 添加客户公司分析数据
        for index, row in full_analysis.iterrows():
            row_num = analysis_sheet.max_row + 1
            
            # 设置单元格值和样式
            cells = [
                (1, row['二级部门'], 'left'),
                (2, row['委托客户'], 'left'),
                (3, row['约价负毛利票数'], 'center'),
                (4, row['约价毛利率'], 'center'),
                (5, row['非约价低负票数'], 'center'),
                (6, row['非约价毛利率'], 'center'),
                (7, row['总票数'], 'center'),
                (8, row['总利润率'], 'center'),
                (9, row['初步分析'], 'left'),
            ]
            
            for col, value, align in cells:
                cell = analysis_sheet.cell(row=row_num, column=col, value=value)
                cell.font = Font(size=9)  # 设置字体大小为9
                cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 设置百分比格式
                if col in [4, 6, 8]:  # 毛利率列
                    cell.number_format = '0.00%'

            # 添加空白列
            for col in range(10, 19):
                cell = analysis_sheet.cell(row=row_num, column=col, value='')
                cell.font = Font(size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

    print(f"分析完成，结果已保存到 {output_file}")
    
    # 在主分析完成后进行拆分
    if status_callback:
        status_callback("正在按部门拆分工作簿...")
    split_workbook_by_department(output_file, business_month)
    if status_callback:
        status_callback("工作簿拆分完成")

def format_analysis(group):
    """
    格式化分析结果，将无应收和倒挂的情况整理成文本描述
    """
    no_receivable = group[group['类型'] == '无应收']['别名'].unique()
    reversed = group[group['类型'] == '倒挂']['别名'].unique()
    
    result = []
    if len(no_receivable) > 0:
        result.append(f"无应收：{', '.join(no_receivable)}")
    if len(reversed) > 0:
        result.append(f"倒挂：{', '.join(reversed)}")
    
    return '\n'.join(result)

def split_workbook_by_department(output_file, business_month):
    """
    将总工作簿按照二级部门和法人部门拆分成多个工作簿
    """
    # 读取原始工作簿中的所有sheet
    all_sheets = pd.read_excel(output_file, sheet_name=None)
    
    # 获取所有二级部门和法人部门
    customer_analysis = pd.read_excel(output_file, sheet_name='客户公司分析', header=[0,1])  # 读取两行表头
    
    # 重命名列，处理多级索引列名
    new_columns = []
    for col in customer_analysis.columns:
        if col[0] == '二级部门' or col[0] == '委托客户' or col[0] == '总票数' or col[0] == '总利润率' or col[0] == '初步分析':
            new_columns.append(col[0])
        elif col[0] == '约价':
            if col[1] == '负毛利票数':
                new_columns.append('约价.负毛利票数')
            elif col[1] == '毛利率':
                new_columns.append('约价.毛利率')
        elif col[0] == '非约价':
            if col[1] == '低负票数':
                new_columns.append('非约价.低负票数')
            elif col[1] == '毛利率':
                new_columns.append('非约价.毛利率')
        else:
            new_columns.append(col[0])
    
    customer_analysis.columns = new_columns
    
    # 获取唯一二级部门
    departments = customer_analysis['二级部门'].unique()
    
    # 从海运订阅原始数据中获取二级部门和法人部门的对应关系
    subscription_data = pd.read_excel(output_file, sheet_name='海运订阅原始数据')
    dept_mapping = {
        '内贸水运': '内贸',
        '外贸水运': '外贸'
    }
    
    # 获取输出文件的目录
    output_dir = os.path.dirname(output_file)
    
    # 为每个部门创建新的工作簿
    for dept in departments:
        # 获取对应的法人部门
        legal_dept = dept_mapping.get(dept, dept)
        
        # 创建新的文件名，使用与总表相同的基础名称，并保持在相同目录
        dept_file = os.path.join(output_dir, f"分析结果_{dept}_{business_month}.xlsx")
        
        try:
            with pd.ExcelWriter(dept_file, engine='openpyxl') as writer:
                # 标记是否有任何数据被写入
                has_data = False
                
                # 处理海运订阅原始数据（按二级部门拆分）
                dept_subscription = subscription_data[subscription_data['二级部门'] == dept]
                if not dept_subscription.empty:
                    dept_subscription.to_excel(writer, sheet_name='海运订阅原始数据', index=False)
                    has_data = True
                
                # 处理预对账原始数据（按法人部门拆分）
                if '预对账原始数据' in all_sheets:
                    precheck_data = all_sheets['预对账原始数据']
                    dept_precheck = precheck_data[precheck_data['法人部门'] == legal_dept]
                    if not dept_precheck.empty:
                        dept_precheck.to_excel(writer, sheet_name='预对账原始数据', index=False)
                        has_data = True
                
                # 处理分析结果（按法人部门拆分）
                if '分析结果' in all_sheets:
                    analysis_data = all_sheets['分析结果']
                    dept_analysis = analysis_data[analysis_data['法人部门'] == legal_dept]
                    if not dept_analysis.empty:
                        dept_analysis.to_excel(writer, sheet_name='分析结果', index=False)
                        has_data = True
                        # 设置分析结果sheet的格式
                        result_sheet = writer.sheets['分析结果']
                        
                        # 设置分析结果sheet的列宽
                        result_widths = {
                            'A': 17,  # 法人部门
                        }
                        
                        # 设置固定列宽
                        for col, width in result_widths.items():
                            result_sheet.column_dimensions[col].width = width
                        
                        # 对其他列进行自适应宽度设置
                        for column in result_sheet.columns:
                            column_letter = get_column_letter(column[0].column)
                            if column_letter not in result_widths:  # 跳过已设置固定宽度的列
                                max_length = 0
                                for cell in column:
                                    try:
                                        if cell.value:
                                            max_length = max(max_length, len(str(cell.value)))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 30)  # 限制最大宽度为30
                                result_sheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # 设置单票毛利率为百分比格式
                        for row in result_sheet.iter_rows(min_row=2):  # 从第2行开始（跳过表头）
                            if row[10].value:  # 第11列是单票毛利率（K列）
                                row[10].number_format = '0.00%'
                
                # 处理客户公司分析（按二级部门拆分）
                dept_customer = customer_analysis[customer_analysis['二级部门'] == dept]
                if not dept_customer.empty:
                    # 创建客户公司分析sheet
                    writer.book.create_sheet('客户公司分析')
                    analysis_sheet = writer.sheets['客户公司分析']
                    
                    # 添加表头
                    headers = [
                        ['二级部门', '委托客户', '约价', '约价', '非约价', '非约价', '总票数', '总利润率', '初步分析', '业务部门反馈具体原因', 
                         '原因类别', '损调利润', '计划采取的措施', '是否完成价格备案表', '是否联合磋商', '督办任务', 
                         '责任人', '督办时间点'],
                        ['', '', '负毛利票数', '毛利率', '低负票数', '毛利率', '', '', '', '', '', '', '', '', '', '', '', '']
                    ]
                    
                    # 设置表头样式
                    for row_index, row in enumerate(headers, start=1):
                        for col_index, value in enumerate(row, start=1):
                            cell = analysis_sheet.cell(row=row_index, column=col_index, value=value)
                            cell.font = Font(bold=True, size=9)
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                    
                    # 合并单元格
                    merge_ranges = [
                        'A1:A2',  # 二级部门
                        'B1:B2',  # 委托客户
                        'C1:D1',  # 约价
                        'E1:F1',  # 非约价
                        'G1:G2',  # 总票数
                        'H1:H2',  # 总利润率
                        'I1:I2',  # 初步分析
                        'J1:J2',  # 业务部门反馈具体原因
                        'K1:K2',  # 原因类别
                        'L1:L2',  # 损调利润
                        'M1:M2',  # 计划采取的措施
                        'N1:N2',  # 是否完成价格备案表
                        'O1:O2',  # 是否联合磋商
                        'P1:P2',  # 督办任务
                        'Q1:Q2',  # 责任人
                        'R1:R2',  # 督办时间点
                    ]
                    
                    # 执行单元格合并
                    for cell_range in merge_ranges:
                        analysis_sheet.merge_cells(cell_range)
                    
                    # 设置表头行高
                    header_height = min(30, 50)  # 表头行高上限为50
                    analysis_sheet.row_dimensions[1].height = header_height
                    analysis_sheet.row_dimensions[2].height = header_height
                    
                    # 写入数据
                    start_row = 3  # 从第3行开始写入数据
                    for _, row in dept_customer.iterrows():
                        try:
                            cells = [
                                (1, row['二级部门'], 'left'),
                                (2, row['委托客户'], 'left'),
                                (3, row['约价.负毛利票数'], 'center'),
                                (4, row['约价.毛利率'], 'center'),
                                (5, row['非约价.低负票数'], 'center'),
                                (6, row['非约价.毛利率'], 'center'),
                                (7, row['总票数'], 'center'),
                                (8, row['总利润率'], 'center'),
                                (9, row['初步分析'], 'left'),
                            ]
                            
                            for col, value, align in cells:
                                cell = analysis_sheet.cell(row=start_row, column=col, value=value)
                                cell.font = Font(size=9)
                                cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
                                cell.border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
                                
                                # 设置百分比格式
                                if col in [4, 6, 8]:  # 毛利率列
                                    if pd.notna(value):  # 只对非空值设置格式
                                        cell.number_format = '0.00%'
                            
                            # 添加空白列
                            for col in range(10, 19):
                                cell = analysis_sheet.cell(row=start_row, column=col, value='')
                                cell.font = Font(size=9)
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
                            
                            start_row += 1
                        except Exception as e:
                            print(f"处理行数据时出错: {str(e)}")
                            continue
                    
                    has_data = True
                    
                    # 设置列宽
                    column_widths = {
                        'A': 15, 'B': 30, 'C': 12, 'D': 10, 'E': 12,
                        'F': 10, 'G': 10, 'H': 10, 'I': 40, 'J': 40,
                        'K': 15, 'L': 12, 'M': 40, 'N': 15, 'O': 15,
                        'P': 15, 'Q': 10, 'R': 15
                    }
                    for col, width in column_widths.items():
                        analysis_sheet.column_dimensions[col].width = width
                    
                    # 设置冻结窗格
                    analysis_sheet.freeze_panes = 'A3'
                
                # 如果没有任何数据被写入，创建一个空的sheet以满足Excel要求
                if not has_data:
                    pd.DataFrame().to_excel(writer, sheet_name='Sheet1', index=False)
                
        except Exception as e:
            print(f"处理部门 {dept} 时出错: {str(e)}")
            continue

if __name__ == "__main__":
    from gui import run_gui
    input_file, output_file, subscription_file = run_gui()
    if subscription_file:
        analyze_excel_data(input_file, output_file, subscription_file)
    else:
        print("未选择海运订阅文件，程序退出")
